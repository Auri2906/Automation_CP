import time
from selenium.webdriver.support.ui import Select
import allure
import openpyxl
import pytest
from allure_commons.types import AttachmentType
from selenium.common.exceptions import NoSuchElementException

from Base import InitializeDriver

global userList
P = input("Enter a names of the profile separated by space ::")
userList = P.split()

driver = InitializeDriver.browser()
driver.get("http://localhost:4200/")


def datagenerator():
    wk = openpyxl.load_workbook("C:\Users\DELL\Desktop\Simnovus\CP_Jenkins\DataSheet\SKYLO.xlsx")
    sh = wk['DEMO3']
    r = sh.max_row
    c = sh.max_column
    li = []
    li2 = []
    for i in range(1, r):
        li1 = []
        for j in range(1, c + 1):
            li1.insert(j - 1, sh.cell(row=i + 1, column=j).value)
        li.insert(i - 1, li1)
    for text in userList:
        if text == '1':
            return li
        else:
            for i in range(0, len(li)):
                if li[i][57] == text.upper():
                    li2.insert(i, li[i])
    return li2


@pytest.mark.parametrize('data', datagenerator())
def test_View_Profile(data):
    c = 0
    s = 0
    u = 0
    t = 0

    try:

        # LOGIN
        driver.find_element_by_id('username').send_keys('user@simnovus.com')
        driver.find_element_by_xpath(
            '/html/body/app-dashboard/div/section/div/div/div[2]/form/div[2]/div/input').send_keys(
            'simnovus')
        driver.find_element_by_xpath(
            '/html/body/app-dashboard/div/section/div/div/div[2]/form/div[3]/div/button').click()

        # VIEW PROFILE
        driver.find_element_by_xpath('/html/body/app-dashboard/div[1]/app-sidebar/div/div[3]/div[1]/a/img').click()
        driver.find_element_by_xpath('//*[@id="manageConfiguration"]').click()

        driver.find_element_by_xpath(
            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div/div[2]/table/thead/tr/th[1]/label/span').click()
        time.sleep(1)
        a = driver.find_element_by_xpath(
            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div/label').text
        n = a.split()
        time.sleep(1)
        driver.find_element_by_xpath(
            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div/div[2]/table/thead/tr/th[1]/label/span').click()
        for i in range(1, int(n[0]) + 1):
            try:
                name = driver.find_element_by_xpath(
                    '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div/div[2]/table/tbody/tr[' + str(
                        i) + ']/td[2]/div').text
                if name.upper() == data[57].upper():
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div/div[2]/table/tbody/tr[' + str(
                            i) + ']/td[1]/label/span').click()
                    break
            except:
                allure.attach("NO profile with name " + P + " in configuration List", name='Profile_NotAvailable')
                assert False
        time.sleep(2)
        with allure.step('View Button and Page title'):
            try:
                driver.find_element_by_xpath(
                    '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div/div[1]/div[2]/span/img[1]').click()
            except NoSuchElementException:
                allure.attach("NO View button is Displayed", name='ViewButton_Error')
            try:
                if driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[1]/span[1]').text == 'Edit Profile':
                    assert True
            except NoSuchElementException:
                allure.attach("Page Title is Displayed Wrong or Missing", name='PageTitle_Error')

        with allure.step('Data validation'):
            with allure.step('Cell Configuration'):
                with allure.step('Profile Name'):
                    try:
                        if driver.find_element_by_xpath(
                                '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[1]/input').get_attribute(
                            'ng-reflect-model') == str(data[57]):
                            assert True

                        else:
                            c += 1
                            allure.attach("Profile Name Mismatch", name='ProfileName_Error')
                    except NoSuchElementException:
                        None
                with allure.step('eNB Type'):
                    try:
                        if data[1] == 4 or data[1] == 44:
                            cell = 'lte'
                        elif data[1] == 444:
                            cell = 'nbiot'
                        elif data[1] == 5:
                            cell = '5G: NSA'
                        else:
                            cell = 'nr'
                        a = Select(driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div[1]/div[1]/div/div/span[2]/select'))
                        b = a.first_selected_option.get_attribute('value')
                        if b == cell:
                            assert True
                        else:
                            c += 1
                            allure.attach("eNB type Mismatch", name='eNBtype_Error')
                    except NoSuchElementException:
                        None
                with allure.step('# Of eNBs'):
                    try:
                        a = Select(driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div[1]/div[2]/div/div/span[2]/select'))
                        b = a.first_selected_option.get_attribute('value')
                        if b == str(data[2]):
                            assert True
                        else:
                            c += 1
                            allure.attach("# Of eNBs Mismatch", name='#OfeNBs_Error')
                    except NoSuchElementException:
                        None
                with allure.step('# Of Cells per eNB'):
                    try:
                        a = driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div[1]/div[3]/div/div/span[2]/select').get_attribute(
                            'ng-reflect-model')
                        if int(a) == data[3]:
                            assert True
                        else:
                            c += 1
                            allure.attach("# Of Cells per eNB Mismatch", name='#OfCells_Error')

                    except NoSuchElementException:
                        None
                with allure.step('DL Antenna'):
                    try:
                        a = driver.find_element_by_xpath(
                            '//*[@id="newDLantennas4g"]').get_attribute(
                            'ng-reflect-model')
                        if int(a) == data[4]:
                            assert True
                        else:
                            c += 1
                            allure.attach("DL Antenna Mismatch", name='DLAntenna_Error')

                    except NoSuchElementException:
                        None
                with allure.step('UL Antenna'):
                    try:
                        a = driver.find_element_by_xpath(
                            '//*[@id="det4g1"]/form/select[4]').get_attribute(
                            'ng-reflect-model')
                        if int(a) == data[5]:
                            assert True
                        else:
                            c += 1
                            allure.attach("UL Antenna Mismatch", name='DLAntenna_Error')

                    except NoSuchElementException:
                        None
                with allure.step('RxGain'):
                    try:
                        a = driver.find_element_by_xpath(
                            '//*[@id="det4g1"]/form/input[2]').get_attribute(
                            'ng-reflect-model')
                        if int(a) == data[8]:
                            assert True
                        else:
                            c += 1
                            allure.attach("RxGain Mismatch", name='RxGain_Error')
                        if data[9] != 'N/A':
                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[3]').get_attribute(
                                'ng-reflect-model')
                            if int(a) == data[9]:
                                assert True
                            else:
                                c += 1
                                allure.attach("RxGain2 Mismatch", name='RxGain2_Error')
                    except NoSuchElementException:
                        None
                with allure.step('TxGain'):
                    try:
                        a = driver.find_element_by_xpath(
                            '//*[@id="det4g1"]/form/input[4]').get_attribute(
                            'ng-reflect-model')
                        if int(a) == data[10]:
                            assert True
                        else:
                            c += 1
                            allure.attach("TxGain Mismatch", name='TxGain_Error')

                    except NoSuchElementException:
                        None
                with allure.step('Duplex Mode'):
                    try:
                        a = driver.find_element_by_xpath(
                            '//*[@id="det4g1"]/form/span[7]/select').get_attribute(
                            'ng-reflect-model')
                        if a == data[11]:
                            assert True
                        else:
                            c += 1
                            allure.attach("Duplex Mode Mismatch", name='DuplexMode_Error')

                    except NoSuchElementException:
                        None
                with allure.step('Band'):
                    try:
                        a = driver.find_element_by_xpath(
                            '//*[@id="bandNew4g"]').get_attribute('value')
                        if a == str(data[12]):
                            assert True
                        else:
                            c += 1
                            allure.attach("Band Mismatch", name='Band_Error')

                    except NoSuchElementException:
                        None

                with allure.step('Bandwidth (MHz)'):
                    try:
                        if data[13] == '180':
                            b = '1.4'
                        a = driver.find_element_by_xpath(
                            '//*[@id="det4g1"]/form/select[2]').get_attribute(
                            'value')
                        if a == str(data[13]) or a == b:
                            assert True
                        else:
                            c += 1
                            allure.attach("Bandwidth (MHz) Mismatch", name='Bandwidth_Error')

                    except NoSuchElementException:
                        None
                '''with allure.step('DL EARFCN'):
                    try:
                        a = driver.find_element_by_xpath(
                            '//*[@id="det4g1"]/form/input[5]').get_attribute(
                            'ng-reflect-model')
                        if data[14] < int(a) < data[15]:
                            assert True
                        else:
                            c += 1
                            allure.attach("DL EARFCN Mismatch", name='DL_EARFCN_Error')
                    except NoSuchElementException:
                        None
                with allure.step('UL EARFCN'):
                    try:
                        a = driver.find_element_by_xpath(
                            '//*[@id="det4g1"]/form/input[6]').get_attribute(
                            'ng-reflect-model')
                        if data[16] < int(a) < data[17]:
                            assert True
                        else:
                            c += 1
                            allure.attach("UL EARFCN Mismatch", name='UL_EARFCN_Error')

                    except NoSuchElementException:
                        None'''
                '''if c > 0:
                    allure.attach(str(c) + ' fields have Wrong or Missing data please check ', name='CELL_CONFIG_ERROR')
                    allure.attach(driver.get_screenshot_as_png(), name='Cell_Config_Screenshot',
                                  attachment_type=AttachmentType.PNG)
                    assert False'''
                driver.find_element_by_xpath('//*[@id="det4g1"]/form/button').click()
                try:
                    if driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div[2]/div[2]/div[1]/div').is_displayed():
                        allure.attach(str(c) + ' fields have Wrong or Missing data please check ',
                                      name='CELL_CONFIG_ERROR')
                        allure.attach(driver.get_screenshot_as_png(), name='Cell_Config_Screenshot',
                                      attachment_type=AttachmentType.PNG)
                        assert False

                except:
                    None

            with allure.step('Subscriber Configuration'):
                with allure.step('Profile Type'):
                    try:
                        if data[19].upper() == 'SINGLE':
                            b = '-1'
                        a = driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[1]/div/div/span[2]/select').get_attribute(
                            'value')
                        if a == b:
                            assert True
                        else:
                            s += 1
                            allure.attach("Profile Type Mismatch", name='Profile_Type_Error')

                    except NoSuchElementException:
                        None
                with allure.step('IMSI'):
                    try:
                        a = driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[2]/div/div/span[2]/input').get_attribute(
                            'value')
                        if a == data[20]:
                            assert True
                        else:
                            s += 1
                            allure.attach("IMSI Mismatch", name='IMSI_Error')

                    except NoSuchElementException:
                        None
                with allure.step('Subscriber '):
                    try:
                        a = driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[2]/div/div/span[3]/input').get_attribute(
                            'value')
                        if int(a) == data[21]:
                            assert True
                        else:
                            s += 1
                            allure.attach("Subscriber e Mismatch", name='Subscriber _Error')

                    except NoSuchElementException:
                        None
                with allure.step('AS Release'):
                    try:
                        a = driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[3]').get_attribute(
                            'value')
                        if int(a) == data[22]:
                            assert True
                        else:
                            s += 1
                            allure.attach("AS Release e Mismatch", name='ASRelease _Error')

                    except NoSuchElementException:
                        None
                with allure.step('IMSI'):
                    try:
                        a = driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/input[2]').get_attribute(
                            'value')
                        if a == data[23]:
                            assert True
                        else:
                            s += 1
                            allure.attach("IMSI Mismatch", name='IMSI_Error')

                    except NoSuchElementException:
                        None
                with allure.step('K'):
                    try:
                        a = driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/input[3]').get_attribute(
                            'value')
                        if a == data[24]:
                            assert True
                        else:
                            s += 1
                            allure.attach("K Mismatch", name='K_Error')

                    except NoSuchElementException:
                        None
                with allure.step('UE Cat Type'):
                    try:
                        if data[25].upper() == 'C':
                            b = '-1'
                        elif data[25].upper == 'D':
                            b = '0'
                        a = driver.find_element_by_xpath(
                            '//*[@id="attachPDNtype"]').get_attribute(
                            'value')
                        if a == b:
                            assert True
                        else:
                            s += 1
                            allure.attach("UE Cat Type Mismatch", name='UECatType_Error')

                    except NoSuchElementException:
                        None

                with allure.step('UE Category '):
                    try:
                        if data[1] != 444:
                            if data[25].upper() == 'D' or data[25].upper() == 'DECOUPLED':
                                a = driver.find_element_by_xpath(
                                    '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[1]').get_attribute(
                                    'value')
                                b = driver.find_element_by_xpath(
                                    '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[2]').get_attribute(
                                    'value')
                                if int(a) == data[27] and int(b) == data[28]:
                                    assert True
                                else:
                                    s += 1
                                    allure.attach("UE Category ", name='UE Category _Error')
                            else:
                                a = driver.find_element_by_xpath(
                                    '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[1]').get_attribute(
                                    'value')
                                if int(a) == data[26]:
                                    assert True
                                else:
                                    s += 1
                                    allure.attach("UE Category ", name='UE Category _Error')
                        else:
                            if data[26] == 'NB1':
                                b = '-2'
                            elif data[26] == 'NB2':
                                b = '-3'
                            else:
                                b = '0'
                            a = driver.find_element_by_xpath(
                                '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[1]').get_attribute(
                                'value')
                            if a == b:
                                assert True
                            else:
                                s += 1
                                allure.attach("UE Category ", name='UE Category _Error')

                    except NoSuchElementException:
                        None
                with allure.step('CQI'):
                    try:
                        if data[30].upper() == 'AUTO':
                            b = 'auto'
                    except:
                        None
                    try:
                        a = driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[7]').get_attribute(
                            'value')
                        if a == b or a == str(data[30]):
                            assert True
                        else:
                            s += 1
                            allure.attach("CQI Mismatch", name='CQI_Error')

                    except NoSuchElementException:
                        None
                with allure.step('Algorithm'):
                    if data[31].upper() == 'M':
                        b = 'milenage'
                    elif data[31].upper() == 'X':
                        b = 'xor'
                    elif data[31].upper() == 'T':
                        b = 'tauk'
                    try:
                        a = driver.find_element_by_xpath(
                            '//*[@id="newSimAlgo"]').get_attribute(
                            'value')
                        if a == b:
                            assert True
                        else:
                            s += 1
                            allure.attach("Algorithm Mismatch", name='Algorithm_Error')

                    except NoSuchElementException:
                        None
                with allure.step('Default APN'):
                    if data[33] == 'N/A':
                        b = ""
                    try:
                        a = driver.find_element_by_xpath(
                            '//*[@id="default_apn"]').get_attribute(
                            'value')
                        if a == data[33] or a == b:
                            assert True
                        else:
                            s += 1
                            allure.attach("Default APN mismatch", name='Default APN_Error')

                    except NoSuchElementException:
                        None
                with allure.step('PDN Type'):
                    if data[34].upper() == 'N/A':
                        b = "ipv4v6"
                    try:
                        a = driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[4]').get_attribute(
                            'value')
                        # d = a.first_selected_option.get_attribute('ng-reflect-value')
                        if a == data[34] or a == b:
                            assert True
                        else:
                            s += 1
                            allure.attach("CQI Mismatch", name='CQI_Error')
                    except NoSuchElementException:
                        None
                with allure.step('Combined Attach'):
                    if data[35] == 'N/A' or data[35].upper() == 'D':
                        b = "-1"
                    else:
                        b = "0"
                    try:
                        a = driver.find_element_by_xpath(
                            '//*[@id="epsIMSIAttach"]').get_attribute(
                            'value')
                        if a == b:
                            assert True
                        else:
                            s += 1
                            allure.attach("Combined Attach mismatch", name='Combined Attach_Error')

                    except NoSuchElementException:
                        None
                with allure.step('SQN'):
                    try:
                        a = driver.find_element_by_xpath(
                            '//*[@id="newSimSQN"]').get_attribute(
                            'value')
                        if a == str(data[36]):
                            assert True
                        else:
                            s += 1
                            allure.attach("SQN mismatch", name='SQN_Error')

                    except NoSuchElementException:
                        None
                with allure.step('RI'):
                    if data[37].upper() == 'AUTO':
                        b = 'auto'
                    try:
                        a = driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[6]').get_attribute(
                            'value')
                        if a == b or a == str(data[37]):
                            assert True
                        else:
                            s += 1
                            allure.attach("RI mismatch", name='RI_Error')

                    except NoSuchElementException:
                        None
                with allure.step('PMI'):
                    if data[38].upper() == 'AUTO':
                        b = 'auto'
                    try:
                        a = driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[6]').get_attribute(
                            'value')
                        if a == b or a == str(data[38]):
                            assert True
                        else:
                            s += 1
                            allure.attach("PMI mismatch", name='PMI_Error')

                    except NoSuchElementException:
                        None

                driver.find_element_by_xpath(
                    '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[1]/label/span').click()
                if data[65] != 'N/A':
                    with allure.step('Half Duplex'):
                        if driver.find_element_by_xpath(
                                '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[11]').is_enabled():
                            if data[65].upper() == 'ENABLE':
                                b = 'true'
                            else:
                                b = 'false'
                            try:
                                a = driver.find_element_by_xpath(
                                    '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[11]').get_attribute(
                                    'value')
                                if a == b:

                                    assert True
                                else:

                                    s += 1
                                    allure.attach("Half Duplex mismatch", name='Half Duplex_Error')

                            except NoSuchElementException:
                                None
                        else:
                            s += 1
                            allure.attach("Half Duplex not enabled", name='Half Duplex_Error')
                if data[66] != 'N/A':
                    with allure.step('Multi Carrier'):
                        if driver.find_element_by_xpath(
                                '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[12]').is_enabled():
                            if data[66].upper() == 'ENABLE':
                                b = 'true'
                            else:
                                b = 'false'
                            try:
                                a = driver.find_element_by_xpath(
                                    '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[12]').get_attribute(
                                    'value')

                                if a == b:

                                    assert True
                                else:

                                    s += 1
                                    allure.attach("Multi Carrier mismatch", name='Multi Carrier_Error')

                            except NoSuchElementException:
                                None
                        else:
                            s += 1
                            allure.attach("Multi Carrier not enabled", name='Multi Carrier_Error')

                if data[67] != 'N/A':
                    with allure.step('Multi Tone'):
                        if driver.find_element_by_xpath(
                                '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[13]').is_enabled():
                            if data[67].upper() == 'ENABLE':
                                b = 'true'
                            else:
                                b = 'false'
                            try:
                                a = driver.find_element_by_xpath(
                                    '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[13]').get_attribute(
                                    'value')

                                if a == b:

                                    assert True
                                else:

                                    s += 1
                                    allure.attach("Multi Tone mismatch", name='Multi Tone_Error')

                            except NoSuchElementException:
                                None
                        else:
                            s += 1
                            allure.attach("Multi Tone not enabled", name='Multi Tone_Error')

                if data[68] != 'N/A':
                    with allure.step('Two HARQ'):
                        if driver.find_element_by_xpath(
                                '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[14]').is_enabled():
                            if data[68].upper() == 'ENABLE':
                                b = '0'
                            else:
                                b = '-1'
                            try:
                                a = driver.find_element_by_xpath(
                                    '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[14]').get_attribute(
                                    'value')

                                if a == b:

                                    assert True
                                else:

                                    s += 1
                                    allure.attach("Two HARQ mismatch", name='Two HARQ_Error')

                            except NoSuchElementException:
                                None
                        else:
                            s += 1
                            allure.attach("Two HARQ not enabled", name='Two HARQ_Error')

                if data[69] != 'N/A':
                    with allure.step('CE Level'):
                        if driver.find_element_by_xpath(
                                '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[15]').is_enabled():
                            try:
                                a = driver.find_element_by_xpath(
                                    '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[15]').get_attribute(
                                    'value')

                                if a == str(data[69]):

                                    assert True
                                else:

                                    s += 1
                                    allure.attach("CE Level mismatch", name='CE Level_Error')

                            except NoSuchElementException:
                                None
                        else:
                            s += 1
                            allure.attach("CE Level not enabled", name='CE Level_Error')

                if data[70] != 'N/A':
                    with allure.step('CP CIoT Opt'):
                        if driver.find_element_by_xpath(
                                '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[7]').is_enabled():
                            if data[70] == True:
                                b = 'true'
                            else:
                                b = 'false'
                            try:
                                a = driver.find_element_by_xpath(
                                    '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/select[7]').get_attribute(
                                    'ng-reflect-model')
                                if a == b:
                                    assert True
                                else:
                                    s += 1
                                    allure.attach("CP CIoT Opt mismatch", name='CP CIoT Opt_Error')

                            except NoSuchElementException:
                                None
                        else:
                            s += 1
                            allure.attach("CP CIoT Opt not enabled", name='CP CIoT Opt_Error')

                '''if s > 0:
                    allure.attach(str(s) + ' fields have Wrong or Missing data please check ',
                                  name='SUBSCRIBER_CONFIG_ERROR')
                    allure.attach(driver.get_screenshot_as_png(), name='SUBSCRIBER_CONFIG_Screenshot',
                                  attachment_type=AttachmentType.PNG)
                    assert False'''
                driver.find_element_by_xpath(
                    '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div[4]/form/div[2]/div[2]/button').click()

            with allure.step('User Plane'):
                with allure.step('Profile Type'):
                    if data[19].upper() == 'SINGLE':
                        b = '-1'
                    try:
                        a = driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div/div[1]/div/div[2]/span[2]/select').get_attribute(
                            'value')
                        if a == b:
                            assert True
                        else:
                            u += 1
                            allure.attach("Profile Type Mismatch", name='profile Type_Error')
                    except NoSuchElementException:
                        None
                if data[40].upper() == 'UDP' or data[40].upper() == 'FTP':
                    with allure.step('Data Type'):
                        if data[40].upper() == 'UDP':
                            b = 'udp'
                        elif data[40].upper() == 'FTP':
                            b = 'ftp'
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/select[1]').get_attribute(
                                'value')
                            if a == b:
                                assert True
                            else:
                                u += 1
                                allure.attach("Data Type Mismatch", name='Data Type_Error')

                        except NoSuchElementException:
                            None
                    with allure.step('Data Direction'):
                        if data[41].upper() == 'DOWNLINK':
                            b = 'downlink'
                        elif data[41].upper() == 'UPLINK':
                            b = 'uplink'
                        else:
                            b = 'both'
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/select[2]').get_attribute(
                                'value')
                            if a == b:
                                assert True
                            else:
                                u += 1
                                allure.attach("Data Direction Mismatch", name='Data Direction_Error')

                        except NoSuchElementException:
                            None
                    if data[41].upper() == 'DOWNLINK':
                        with allure.step('DL Bitrate'):
                            try:

                                a = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/input[1]').get_attribute(
                                    'value')
                                if a == str(data[42]) or a == str(data[42]/1000):
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("DL Bitrate Mismatch", name='DL Bitrate_Error')

                            except NoSuchElementException:
                                None
                    elif data[41].upper() == 'UPLINK':
                        with allure.step('UL Bitrate'):
                            try:
                                a = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/input[2]').get_attribute(
                                    'value')
                                if a == str(data[43]) or a == str(data[43]/1000):
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("UL Bitrate Mismatch", name='UL Bitrate_Error')

                            except NoSuchElementException:
                                None
                    else:
                        with allure.step('DL Bitrate'):
                            try:

                                a = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/input[1]').get_attribute(
                                    'value')
                                if a == str(data[42]) or a == str(data[42]/1000):
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("DL Bitrate Mismatch", name='DL Bitrate_Error')

                            except NoSuchElementException:
                                None
                        with allure.step('UL Bitrate'):
                            try:
                                a = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/input[2]').get_attribute(
                                    'value')
                                if a == str(data[43]) or a == str(data[43]/1000):
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("UL Bitrate Mismatch", name='UL Bitrate_Error')

                            except NoSuchElementException:
                                None
                    with allure.step('Data Pattern'):
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/select[5]').get_attribute(
                                'value')
                            if a.upper() == str(data[44]).upper():
                                assert True
                            else:
                                u += 1
                                allure.attach("Data Pattern Mismatch", name='Data Pattern_Error')

                        except NoSuchElementException:
                            None
                    if driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/input[3]').is_enabled():
                        with allure.step('APN Name'):
                            try:
                                a = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/input[3]').get_attribute(
                                    'value')
                                if a == data[45]:
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("APN Name Mismatch", name='APN Name_Error')

                            except NoSuchElementException:
                                None
                    with allure.step('IP Address'):
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/input[5]').get_attribute(
                                'ng-reflect-model')
                            if a == str(data[46]):
                                assert True
                            else:
                                u += 1
                                allure.attach("IP Address Mismatch", name='IP Address_Error')

                        except NoSuchElementException:
                            None
                    with allure.step('Start Delay(secs)'):
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/input[6]').get_attribute(
                                'value')
                            if a == str(data[47]):
                                assert True
                            else:
                                u += 1
                                allure.attach("Start Delay(secs) Mismatch", name='IP Address_Error')

                        except NoSuchElementException:
                            None
                    with allure.step('Duration (secs)'):
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/input[7]').get_attribute(
                                'value')
                            if a == str(data[48]):
                                assert True
                            else:
                                u += 1
                                allure.attach("Duration Mismatch", name='Duration_Error')

                        except NoSuchElementException:
                            None
                    if data[44] == 'Intermittent':
                        with allure.step('InterPacket Delay'):
                            try:
                                a = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/input[10]').get_attribute(
                                    'value')
                                if a == str(data[49]):
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("InterPacket Delay Mismatch", name='InterPacket Delay_Error')

                            except NoSuchElementException:
                                None
                    if data[40].upper() == 'UDP':
                        with allure.step('Payload Length(byte)'):
                            try:
                                a = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/input[4]').get_attribute(
                                    'value')
                                if a == str(data[50]):
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("Payload Length(byte) Mismatch", name='Payload Length(byte)_Error')

                            except NoSuchElementException:
                                None
                    if data[51] != 'N/A':
                        with allure.step('IMSI Start'):
                            try:
                                a = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/input[9]').get_attribute(
                                    'value')
                                if a == str(data[51]):
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("IMSI Start) Mismatch", name='IMSI Start_Error')

                            except NoSuchElementException:
                                None
                    with allure.step('# of Subscribers'):
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/input[8]').get_attribute(
                                'value')
                            if a == str(data[21]):
                                assert True
                            else:
                                u += 1
                                allure.attach("# of Subscribers Start) Mismatch", name='# of Subscribers_Error')

                        except NoSuchElementException:
                            None
                    if data[40].upper() == 'FTP':
                        with allure.step('Password'):
                            try:
                                a = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/div/input').get_attribute(
                                    'value')
                                if a == str(data[71]):
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("Password Mismatch", name='Password_Error')

                            except NoSuchElementException:
                                None

                elif data[40].upper() == 'VOLTE':
                    with allure.step('Data Type'):
                        if data[40].upper() == 'VOLTE':
                            b = 'volte'
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/select[1]').get_attribute(
                                'value')
                            if a == b:
                                assert True
                            else:
                                u += 1
                                allure.attach("Data Type Mismatch", name='Data Type_Error')

                        except NoSuchElementException:
                            None

                    with allure.step('APN Name'):
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/input[3]').get_attribute(
                                'value')
                            if a == data[45]:
                                assert True
                            else:
                                u += 1
                                allure.attach("APN Name Mismatch", name='APN Name_Error')
                        except NoSuchElementException:
                            None

                    with allure.step('Start Delay(secs)'):
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/input[7]').get_attribute(
                                'value')
                            if a == str(data[47]):
                                assert True
                            else:
                                u += 1
                                allure.attach("Start Delay(secs) Mismatch", name='Start Delay(secs)_Error')

                        except NoSuchElementException:
                            None
                    with allure.step('Duration (secs)'):
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/input[8]').get_attribute(
                                'value')
                            if a == str(data[48] + data[47]):
                                assert True
                            else:
                                u += 1
                                allure.attach("Duration Mismatch", name='Duration_Error')

                        except NoSuchElementException:
                            None
                    if data[51] != 'N/A':
                        with allure.step('IMSI Start'):
                            try:
                                a = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/input[9]').get_attribute(
                                    'value')
                                if a == str(data[51]):
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("IMSI Start) Mismatch", name='IMSI Start_Error')

                            except NoSuchElementException:
                                None
                    with allure.step('# of Subscribers'):
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/input[9]').get_attribute(
                                'value')
                            if a == str(data[21]):
                                assert True
                            else:
                                u += 1
                                allure.attach("# of Subscribers Start) Mismatch", name='# of Subscribers_Error')

                        except NoSuchElementException:
                            None
                    with allure.step('IMS Profile'):
                        with allure.step('Selected IMS Profile'):
                            try:
                                a = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/input[6]').get_attribute(
                                    'value')
                                if a == data[57]:
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("Selected IMS profile mismatch", name='Selected_IMS_Profile_Error')

                            except NoSuchElementException:
                                None
                        #driver.find_element_by_xpath('').click()
                        with allure.step('Profile Name'):
                            try:
                                a = driver.find_element_by_xpath(
                                    '//*[@id="myModal"]/div/div/div[2]/input[1]').get_attribute(
                                    'value')
                                if a == data[57]:
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("profile mismatch", name='Profile_name_Error')

                            except NoSuchElementException:
                                None
                        with allure.step('Pcscf Ip'):
                            try:
                                a = driver.find_element_by_xpath(
                                    '//*[@id="myModal"]/div/div/div[2]/input[2]').get_attribute(
                                    'value')
                                if a == data[59]:
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("Pcscf Ip mismatch", name='Pcscf Ip_Error')

                            except NoSuchElementException:
                                None
                        '''with allure.step('Algorithm'):
                            try:
                                if data[60].upper() == 'AKA DIGEST':
                                    b = '-1'
                                a = driver.find_element_by_xpath(
                                    '//*[@id="myModal"]/div/div/div[2]/select[1]').get_attribute(
                                    'value')
                                if a == b:
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("Algorithm mismatch", name='Algorithm_Error')

                            except NoSuchElementException:
                                None'''
                        with allure.step('Password'):
                            try:
                                a = driver.find_element_by_xpath(
                                    '//*[@id="myModal"]/div/div/div[2]/input[4]').get_attribute(
                                    'value')
                                if a == str(data[61]):
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("Password mismatch", name='Password_Error')
                            except NoSuchElementException:
                                None
                        with allure.step('Codec'):
                            try:
                                a = driver.find_element_by_xpath(
                                    '//*[@id="myModal"]/div/div/div[2]/select[2]').get_attribute(
                                    'value')
                                if a == data[62]:
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("Codec mismatch", name='Codec_Error')
                            except NoSuchElementException:
                                None
                        with allure.step('Sampling Rate'):
                            try:
                                a = driver.find_element_by_xpath(
                                    '//*[@id="myModal"]/div/div/div[2]/input[5]').get_attribute(
                                    'value')
                                if a == str(data[63]):
                                    assert True
                                else:
                                    u += 1
                                    allure.attach("Sampling Rate mismatch", name='Sampling Rate_Error')
                            except NoSuchElementException:
                                None

                '''if u > 0:
                    allure.attach(str(u) + ' fields have Wrong or Missing data please check ',
                                  name='SUBSCRIBER_CONFIG_ERROR')
                    allure.attach(driver.get_screenshot_as_png(), name='SUBSCRIBER_CONFIG_Screenshot',
                                  attachment_type=AttachmentType.PNG)
                    assert False'''
                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/button').click()

            with allure.step('Traffic Profile'):
                with allure.step('Profile Type'):
                    if data[19].upper() == 'SINGLE':
                        b = '-1'
                    try:
                        a = driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div[2]/div/div/span[2]/select').get_attribute(
                            'value')
                        if a == b:
                            assert True
                        else:
                            t += 1
                            allure.attach("Profile Type Mismatch", name='profile Type_Error')
                    except NoSuchElementException:
                        None
                if data[53].upper() == 'STAGGERED':
                    with allure.step('Attach Type'):
                        if data[53].upper() == 'STAGGERED':
                            b = 'Staggered'
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/select').get_attribute(
                                'value')
                            if a == b:
                                assert True
                            else:
                                t += 1
                                allure.attach("Attach Type Mismatch", name='Attach Type_Error')
                        except NoSuchElementException:
                            None

                    with allure.step('Attach Rate'):
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/input[1]').get_attribute(
                                'value')
                            if a == str(data[54]):
                                assert True
                            else:
                                t += 1
                                allure.attach("Attach Rate Mismatch", name='Attach Rate_Error')
                        except NoSuchElementException:
                            None
                    with allure.step('Power On Time'):
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/input[2]').get_attribute(
                                'value')
                            if a == str(data[55]):
                                assert True
                            else:
                                t += 1
                                allure.attach("Power On Time Mismatch", name='Attach Rate_Error')
                        except NoSuchElementException:
                            None
                    with allure.step('Power off Time'):
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/input[3]').get_attribute(
                                'value')
                            if a == str(data[56]):
                                assert True
                            else:
                                t += 1
                                allure.attach("Power 0ff Time Mismatch", name='Attach Rate_Error')
                        except NoSuchElementException:
                            None
                else:
                    with allure.step('Attach Type'):
                        if data[53].upper() == 'BURSTY':
                            b = 'Bursty'
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/select').get_attribute(
                                'value')
                            if a == b:
                                assert True
                            else:
                                t += 1
                                allure.attach("Attach Type Mismatch", name='Attach Type_Error')
                        except NoSuchElementException:
                            None

                    with allure.step('Power On Time'):
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/input[2]').get_attribute(
                                'value')
                            if a == str(data[55]):
                                assert True
                            else:
                                t += 1
                                allure.attach("Power On Time Mismatch", name='Attach Rate_Error')
                        except NoSuchElementException:
                            None
                    with allure.step('Power off Time'):
                        try:
                            a = driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/input[3]').get_attribute(
                                'value')
                            if a == str(data[56]):
                                assert True
                            else:
                                t += 1
                                allure.attach("Power 0ff Time Mismatch", name='Attach Rate_Error')
                        except NoSuchElementException:
                            None
                '''if t > 0:
                    allure.attach(str(t) + ' fields have Wrong or Missing data please check ',
                                  name='SUBSCRIBER_CONFIG_ERROR')
                    allure.attach(driver.get_screenshot_as_png(), name='SUBSCRIBER_CONFIG_Screenshot',
                                  attachment_type=AttachmentType.PNG)
                    assert False'''
                with allure.step('Update Button'):
                    try:
                        driver.find_element_by_xpath(
                            '//*[@id="det4g1"]/form/div/div/div[1]/div/button[1]').is_displayed()

                    except NoSuchElementException:
                        allure.attach("Update Button not Displayed", name='UpdateButton_Error')
                    try:
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/div[1]/div/button[1]').is_enabled()

                    except NoSuchElementException:
                        allure.attach("Update Button not Enabled", name='UpdateButton_Error')

                with allure.step('Cancel Button'):
                    try:
                        driver.find_element_by_xpath(
                            '//*[@id="det4g1"]/form/div/div/div[1]/div/button[2]').is_displayed()
                    except NoSuchElementException:
                        allure.attach("Cancel Button not Displayed", name='CancelButton_Error')
                    try:
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/div[1]/div/button[2]').click()
                        time.sleep(2)
                        if driver.find_element_by_xpath(
                                '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div/div[1]/div[1]/span').is_displayed():
                            assert True
                        else:
                            allure.attach("Cancel Button not Enabled", name='CancelButton_Error')
                    except NoSuchElementException:
                        None


    finally:
        driver.find_element_by_xpath('/html/body/app-dashboard/app-header/ul/li[2]/a/img').click()
        driver.find_element_by_xpath('/html/body/app-dashboard/app-header/ul/li[2]/div/a[4]').click()