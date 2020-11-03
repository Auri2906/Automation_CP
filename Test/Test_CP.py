import allure
from allure_commons.types import AttachmentType
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import Select
import random
import openpyxl
import pytest
import time
from Base import InitializeDriver

print("please provide the input for eNB type in the following format:: ")
print("4   :: 4G: Smartphone")
print("44  :: 4G: Cat-M")
print("444 :: 4G: Nb-IoT")
print("5   :: 5G: NSA")
print("55  :: 5G: SA")

global userList
eNB = input("Enter a list eNB separated by space ::")
print("\n")
userList = eNB.split()

sdr = InitializeDriver.sdr()
driver = InitializeDriver.browser()


def datagenerator():
    wk = openpyxl.load_workbook("C:/Users/DELL/Desktop/Simnovus/CP_Jenkins/DataSheet/SKYLO.xlsx")
    sh = wk['DEMO3']
    r = sh.max_row
    c = sh.max_column
    li = []
    li2 = []
    for i in range(1, r):
        li1 = []
        for j in range(1, c + 1):
            li1.insert(j - 1, sh.cell(row=i + 1, column=j).value)
        li.insert(i - 1, li1)
    for text in userList:
        if int(text) == 1:
            return li
        elif int(text) == 4 or int(text) == 44 or int(text) == 444 or int(text) == 5 or int(text) == 55:
            for i in range(0, len(li)):
                if li[i][1] == int(text):
                    li2.insert(i, li[i])
        else:
            print('Wrong Input!!!')
            print('All tests Skipped')
            print('Run test again !!!')
            print("\n")
    return li2


@pytest.mark.parametrize('data', datagenerator())
def test_config_1(data):
    try:
        z = 0
        R = 0
        CELL = 0
        SUB = 0
        USER = 0
        global driver
        try:

            global A
            A = False
            with allure.step("Launch Webpage"):
                driver.get("http://localhost:4200/")
                driver.implicitly_wait(10)
            with allure.step("Login to Simnovator GUI"):
                driver.find_element_by_id('username').send_keys('user@simnovus.com')
                driver.find_element_by_xpath(
                    '/html/body/app-dashboard/div/section/div/div/div[2]/form/div[2]/div/input').send_keys(
                    'simnovus')

                driver.find_element_by_xpath(
                    '/html/body/app-dashboard/div/section/div/div/div[2]/form/div[3]/div/button').click()

                try:
                    if driver.find_element_by_xpath(
                        "/html/body/app-dashboard/div[1]/main/div/app-welcome/div/div[1]/span[1]").is_displayed():
                        assert True


                except NoSuchElementException:
                    if driver.find_element_by_xpath("/html/body/app-dashboard/div/div[2]/h1").is_displayed():
                        allure.attach(driver.get_screenshot_as_png(), name='PM2_Server_failure',
                                      attachment_type=AttachmentType.PNG)
                    assert False

                A = False
                B = False
                try:
                    A = driver.find_element_by_xpath(
                        "/html/body/app-dashboard/div[1]/main/div/app-welcome/div/div[1]/span[1]").is_displayed()
                except:
                    B = driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div/section/div/div/div[2]/form/div[2]/div[2]').is_displayed()
                finally:
                    if A or B:
                        assert True
                    else:
                        allure.attach(driver.get_screenshot_as_png(), name='TestLogin_failure',
                                      attachment_type=AttachmentType.PNG)
                        assert False

            with allure.step('CELL Configuration'):
                # assert driver.current_url == 'http://192.168.0.104/#/login'
                driver.find_element_by_xpath(
                    '/html/body/app-dashboard/div[1]/app-sidebar/div/div[3]/div[1]/a/img').click()
                driver.find_element_by_xpath('//*[@id="newConfiguration"]').click()

                # assert driver.current_url == 'http://192.168.0.104/#/welcome'
                eNBTYPE = ['Select', '4G: Smartphone ', '4G: Cat-M', '4G: Nb-IoT', '5G: NSA', '5G: SA', 'Reset']
                dropdown = Select(driver.find_element_by_xpath(
                    '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div[1]/div[1]/div/div/span[2]/select'))
                i = 0
                c = 0
                li = []
                for element in dropdown.options:
                    li.insert(i, element.text)
                    i += 1
                for i in range(0, len(eNBTYPE)):
                    if eNBTYPE.__getitem__(i) in li:
                        assert True
                    else:
                        c += 1
                if c != 0:
                    allure.attach("Dropdown Values of 'eNB Type' not as Expected", name='DropdownError_eNBType')

                driver.find_element_by_css_selector(".setcellstatusbox4g").click()
                driver.find_element_by_css_selector(".setcellstatusbox4g").send_keys(data[1])
                driver.find_element_by_css_selector(".setcellstatusbox4g").click()

                time.sleep(1)
                if sdr == 1:
                    OfeNBs = ['Select', '1']
                elif sdr == 2:
                    OfeNBs = ['Select', '1', '2']
                elif sdr == 3:
                    OfeNBs = ['Select', '1', '2', '3']
                elif sdr == 4:
                    OfeNBs = ['Select', '1', '2', '3', '4']
                elif sdr == 5:
                    OfeNBs = ['Select', '1', '2', '3', '4', '5']
                elif sdr == 6:
                    OfeNBs = ['Select', '1', '2', '3', '4', '5', '6']
                else:
                    OfeNBs = []

                dropdown = Select(driver.find_element_by_xpath(
                    '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[2]/div/div/span[2]/select'))
                i = 0
                c = 0
                li = []
                if OfeNBs == []:
                    for element in dropdown.options:
                        OfeNBs.insert(i, '0')
                        li.insert(i, element.text)
                        i += 1
                else:
                    for element in dropdown.options:
                        li.insert(i, element.text)
                        i += 1

                for i in range(0, len(OfeNBs)):
                    if OfeNBs.__getitem__(i) in li:
                        assert True
                    else:
                        c += 1
                if c != 0:
                    allure.attach("Dropdown Values of '# of eNBS' not as Expected", name='DropdownError_#ofeNBs')

                driver.find_element_by_css_selector(".setcellstatusbox4g").click()
                driver.find_element_by_xpath(
                    "/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div["
                    "2]/div/div[2]/div/div/span[2]/select").click()
                driver.find_element_by_xpath(
                    "/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div["
                    "2]/div/div[2]/div/div/span[2]/select").send_keys(data[2])
                driver.find_element_by_xpath(
                    "/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div["
                    "2]/div/div[2]/div/div/span[2]/select").click()
                time.sleep(1)
                if sdr == 1:
                    Ofcell = ['Select', '1']
                elif sdr == 2:
                    Ofcell = ['Select', '1', '2']
                elif sdr == 3:
                    Ofcell = ['Select', '1', '2', '3']
                elif sdr == 4:
                    Ofcell = ['Select', '1', '2', '3', '4']
                elif sdr == 5:
                    Ofcell = ['Select', '1', '2', '3', '4', '5']
                elif sdr == 6:
                    Ofcell = ['Select', '1', '2', '3', '4', '5', '6']
                else:
                    Ofcell = []
                dropdown2 = Select(driver.find_element_by_xpath(
                    '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[3]/div/div/span[2]/select'))
                i = 0
                c = 0
                li2 = []
                if Ofcell == []:
                    for element in dropdown2.options:
                        Ofcell.insert(i, '0')
                        li2.insert(i, element.text)
                        i += 1
                else:
                    for element in dropdown2.options:
                        li2.insert(i, element.text)
                        i += 1
                for i in range(0, len(Ofcell)):
                    if Ofcell.__getitem__(i) in li2:
                        assert True
                    else:
                        c += 1
                if c != 0:
                    allure.attach("Dropdown Values of '# of Cells' not as Expected", name='DropdownError_#ofCells')

                driver.find_element_by_xpath(
                    "/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div["
                    "2]/div/div[3]/div/div/span[2]/select").click()
                driver.find_element_by_xpath(
                    "/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div["
                    "2]/div/div[3]/div/div/span[2]/select").send_keys(data[3])
                driver.find_element_by_xpath(
                    "/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div["
                    "2]/div/div[3]/div/div/span[2]/select").click()

                if data[1] == 'R':
                    if driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div[2]/button').is_displayed():
                        allure.attach(driver.get_screenshot_as_png(), name='TestCellConfig_Failure',
                                      attachment_type=AttachmentType.PNG)
                        R = 1
                        assert False

                # time.sleep(1)

                if data[1] == 5 or data[1] == 55:

                    driver.find_element_by_id("newDLantennas4g").click()
                    driver.find_element_by_xpath("//*[@id='newDLantennas4g']").send_keys(data[4])
                    driver.find_element_by_xpath('//*[@id="newULantennas4g"]').click()
                    driver.find_element_by_xpath('//*[@id="newULantennas4g"]').send_keys(data[5])

                    # time.sleep(1)

                    driver.find_element_by_xpath(
                        "//html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div["
                        "2]/div/div/form/input[3]").click()
                    if data[6] == 'tick':
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/select[1]').click()
                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/select[1]').send_keys(data[6])

                    # time.sleep(1)

                    driver.find_element_by_id("bandNew4g").click()
                    driver.find_element_by_xpath("//*[@id='bandNew4g']").send_keys(data[7])

                    # time.sleep(1)

                    driver.find_element_by_xpath("//*[@id='det4g1']/form/select[3]").click()
                    driver.find_element_by_xpath("//*[@id='det4g1']/form/select[3]").send_keys(data[8])

                    # time.sleep(1)

                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[1]').click()
                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[1]').clear()
                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[1]').send_keys("")
                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[1]').send_keys(
                        random.randint(data[8], data[9]))

                    # time.sleep(1)

                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[2]').click()
                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[2]').clear()
                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[2]').send_keys("")
                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[2]').send_keys(
                        random.randint(data[10], data[11]))

                    # time.sleep(1)

                    driver.find_element_by_id('newBandwidth4g').click()
                    driver.find_element_by_id('newBandwidth4g').send_keys(data[12])
                    driver.find_element_by_id('newBandwidth4g').click()

                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/button').click()

                else:
                    CellType = ['4G']
                    dropdown3 = Select(driver.find_element_by_xpath(
                        '//*[@id="det4g1"]/form/span[4]/select'))
                    i = 0
                    c = 0
                    li3 = []
                    for element in dropdown3.options:
                        li3.insert(i, element.text)
                        i += 1
                    for i in range(0, len(CellType)):
                        if CellType.__getitem__(i) in li3:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'Cell Type' not as Expected", name='DropdownError_CellType')

                    if data[2] == 1:
                        eNB = ['1']
                    elif data[2] == 2:
                        eNB = ['1', '2']
                    elif data[2] == 3:
                        eNB = ['1', '2', '3']
                    elif data[2] == 4:
                        eNB = ['1', '2', '3', '4']
                    elif data[2] == 5:
                        eNB = ['1', '2', '3', '4', '5']
                    elif data[2] == 6:
                        eNB = ['1', '2', '3', '4', '5', '6']
                    else:
                        eNB = []
                    dropdown4 = Select(driver.find_element_by_xpath(
                        '//*[@id="det4g1"]/form/span[10]/select'))
                    i = 0
                    c = 0
                    li4 = []
                    if eNB == []:
                        for element in dropdown4.options:
                            eNB.insert(i, '0')
                            li4.insert(i, element.text)
                            i += 1
                    else:
                        for element in dropdown4.options:
                            li4.insert(i, element.text)
                            i += 1
                    for i in range(0, len(eNB)):
                        if eNB.__getitem__(i) in li4:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'eNBS #' not as Expected", name='DropdownError_eNB#')

                    if sdr == 1:
                        DL = ['Select', '1', '2']
                    elif sdr == 2:
                        DL = ['Select', '1', '2', '3', '4']
                    elif sdr == 3:
                        DL = ['Select', '1', '2', '3', '4', '5', '6']
                    elif sdr == 4:
                        DL = ['Select', '1', '2', '3', '4', '5', '6', '7', '8']
                    elif sdr == 5:
                        DL = ['Select', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10']
                    elif sdr == 6:
                        DL = ['Select', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']
                    else:
                        DL = []
                    dropdown5 = Select(driver.find_element_by_xpath(
                        '//*[@id="newDLantennas4g"]'))
                    i = 0
                    c = 0
                    li5 = []
                    if DL == []:
                        for element in dropdown5.options:
                            DL.insert(i, '0')
                            li5.insert(i, element.text)
                            i += 1
                    else:
                        for element in dropdown5.options:
                            li5.insert(i, element.text)
                            i += 1

                    for i in range(0, len(DL)):
                        if DL.__getitem__(i) in li5:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'DL Antennas' is not as Expected",
                                      name='DropdownError_DLAntenna')

                    driver.find_element_by_xpath("//*[@id='newDLantennas4g']").click()
                    driver.find_element_by_xpath("//*[@id='newDLantennas4g']").send_keys(data[4])
                    driver.find_element_by_xpath("//*[@id='newDLantennas4g']").click()

                    UL = ['Select', '1']
                    dropdown6 = Select(driver.find_element_by_xpath('//*[@id="det4g1"]/form/select[4]'))
                    i = 0
                    c = 0
                    li6 = []
                    for element in dropdown6.options:
                        li6.insert(i, element.text)
                        i += 1
                    for i in range(0, len(UL)):
                        if UL.__getitem__(i) in li6:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'UL Antennas' is not as Expected",
                                      name='DropdownError_ULAntenna')

                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/select[4]').click()
                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/select[4]').send_keys(data[5])
                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/select[4]').click()

                    # time.sleep(2)
                    if data[6] == 'tick':
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[1]').click()
                        time.sleep(1)

                        if data[1] == 44:
                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[2]').click()
                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[2]').clear()
                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[2]').send_keys(data[8])

                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[3]').click()
                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[3]').clear()
                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[3]').send_keys(data[10])

                        else:

                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[2]').click()
                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[2]').clear()
                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[2]').click()
                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[2]').send_keys(data[8])

                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[3]').click()
                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[3]').clear()
                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[3]').click()
                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[3]').send_keys(data[10])
                    if data[1] == 44:
                        DM = ['FDD']
                    else:
                        DM = ['FDD', 'TDD']
                    dropdown6 = Select(driver.find_element_by_xpath('//*[@id="det4g1"]/form/span[7]/select'))
                    i = 0
                    c = 0
                    li6 = []
                    for element in dropdown6.options:
                        li6.insert(i, element.text)
                        i += 1
                    for i in range(0, len(DM)):
                        if DM.__getitem__(i) in li6:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'Duplex Mode' is not as Expected",
                                      name='DropdownError_DuplexType')

                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/span[7]/select').click()
                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/span[7]/select').send_keys(data[11])

                    # time.sleep(1)

                    driver.find_element_by_xpath('//*[@id="bandNew4g"]').click()
                    if data[11] == 'FDD':
                        Band = ['Select', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14',
                                '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '30', '31',
                                '65', '66', '68', '70', '71', '72', '73', '74', '85']
                    else:
                        Band = ['Select', '33', '34', '35', '36', '37', '38', '39', '40', '41', '42', '43', '44', '45',
                                '46', '47', '48', '49', '50', '51', '52']
                    dropdown6 = Select(driver.find_element_by_xpath('//*[@id="bandNew4g"]'))
                    i = 0
                    li6 = []
                    c = 0

                    for element in dropdown6.options:
                        li6.insert(i, element.text)
                        i += 1
                    for i in range(0, len(Band)):
                        if Band.__getitem__(i) in li6:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'Band' is not as Expected", name='DropdownError_Band')

                    driver.find_element_by_id("bandNew4g").click()
                    driver.find_element_by_xpath("//*[@id='bandNew4g']").send_keys(data[12])
                    driver.find_element_by_id("bandNew4g").click()

                    # time.sleep(1)

                    if data[12] == 1 or data[12] == 7 or data[12] == 9 or data[12] == 10 or data[12] == 20 or data[
                        12] == 22 or data[12] == 33 or data[12] == 37 or data[12] == 38 or data[12] == 39 or data[
                        12] == 40 or data[12] == 41 or data[12] == 42 or data[12] == 43 or data[12] == 45 or data[
                        12] == 48 or data[12] == 52 or data[12] == 70 or data[12] == 71:
                        BW = ['Select value', '5', '10', '15', '20']
                    elif data[12] == 2 or data[12] == 3 or data[12] == 4 or data[12] == 23 or data[12] == 25 or data[
                        12] == 35 or data[12] == 36 or data[12] == 65 or data[12] == 66 or data[12] == 74:
                        BW = ['Select value', '1.4', '3', '5', '10', '15', '20']
                    elif data[12] == 5 or data[12] == 8 or data[12] == 12 or data[12] == 27:
                        BW = ['Select value', '1.4', '3', '5', '10']
                    elif data[12] == 6 or data[12] == 11 or data[12] == 13 or data[12] == 14 or data[12] == 17 or data[
                        12] == 24 or data[12] == 30 or data[12] == 85:
                        BW = ['Select value', '5', '10']
                    elif data[12] == 18 or data[12] == 19 or data[12] == 21 or data[12] == 34 or data[12] == 68:
                        BW = ['Select value', '5', '10', '15']
                    elif data[12] == 26:
                        BW = ['Select value', '1.4', '3', '5', '10', '15']
                    elif data[12] == 28 or data[12] == 44 or data[12] == 50:
                        BW = ['Select value', '3', '5', '10', '15', '20']
                    elif data[12] == 31 or data[12] == 72 or data[12] == 73:
                        BW = ['Select value', '1.4', '3', '5']
                    elif data[12] == 46 or data[12] == 47 or data[12] == 49:
                        BW = ['Select value', '10', '20']
                    elif data[12] == 28 or data[12] == 44 or data[12] == 50:
                        BW = ['Select value', '3', '5', '10', '15', '20']
                    else:
                        BW = ['Select value', '3', '5']
                    dropdown6 = Select(driver.find_element_by_xpath('//*[@id="det4g1"]/form/select[2]'))
                    i = 0
                    c = 0
                    li6 = []
                    for element in dropdown6.options:
                        li6.insert(i, element.text)
                        i += 1
                    for i in range(0, len(BW)):
                        if BW.__getitem__(i) in li6:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'BandWidth(MHz)' is not as Expected",
                                      name='DropdownError_BandWidth(MHz)')

                    driver.find_element_by_xpath("//*[@id='det4g1']/form/select[2]").click()
                    driver.find_element_by_xpath("//*[@id='det4g1']/form/select[2]").send_keys(data[13])
                    driver.find_element_by_xpath("//*[@id='det4g1']/form/select[2]").click()

                    # time.sleep(1)

                    '''if data[1] == 44:
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[4]').click()
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[4]').clear()
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[4]').send_keys("")
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[4]').send_keys(
                            random.randint(data[14], data[15]))

                        # time.sleep(1)

                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[5]').click()
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[5]').clear()
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[5]').send_keys("")
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[5]').send_keys(
                            random.randint(data[16], data[17]))


                    else:
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[4]').click()
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[4]').clear()
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[4]').send_keys("")
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[4]').send_keys(
                            random.randint(data[14], data[15]))

                        # time.sleep(1)

                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[5]').click()
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[5]').clear()
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[5]').send_keys("")
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/input[5]').send_keys(
                            random.randint(data[16], data[17]))'''

                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/button').click()

                    try:
                        A = driver.find_element_by_xpath('/html/body/app-dashboard/div['
                                                         '1]/main/div/app-create-configuration/div/div[2]/span').is_displayed()
                        B = driver.find_element_by_xpath('/html/body/app-dashboard/div['
                                                         '1]/main/div/app-create-configuration/div/div[2]/div[2]/div[2]/div['
                                                         '1]/div').is_displayed()
                    except:
                        None

                    finally:
                        if A or B:
                            assert True
                            if B:
                                CELL = 1
                        else:
                            allure.attach(driver.get_screenshot_as_png(), name='Testcell_Failure',
                                          attachment_type=AttachmentType.PNG)
                            assert False


        except:
            if R == 1:
                assert False
            else:
                driver.find_element_by_xpath('//*[@id="det4g1"]/form/button').click()
                allure.attach(driver.get_screenshot_as_png(), name='Testcell_missingdata',
                              attachment_type=AttachmentType.PNG)
                if driver.find_element_by_xpath('/html/body/app-dashboard/div['
                                                '1]/main/div/app-create-configuration/div/div[2]/div[2]/div[2]/div['
                                                '1]/div').is_displayed():
                    CELL = 1
                    assert True
                else:
                    allure.attach(driver.get_screenshot_as_png(), name='Testcell_Failure',
                                  attachment_type=AttachmentType.PNG)
                    assert False

        if CELL == 0:
            try:
                with allure.step("Subscriber Configuration"):

                    PT = ['Type', 'Single Profile', 'Mixed Profile', 'Reset']

                    dropdown7 = Select(driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div/div/div/span[2]/select'))
                    i = 0
                    c = 0
                    li7 = []
                    for element in dropdown7.options:
                        li7.insert(i, element.text)
                        i += 1
                    for i in range(0, len(PT)):
                        if PT.__getitem__(i) in li7:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'Profile Type' is not as Expected",
                                      name='DropdownError_ProfileType')

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                        '2]/div/div/div/div/span[2]/select').click()

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                        '2]/div/div/div/div/span[2]/select').send_keys(data[19])
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                        '2]/div/div/div/div/span[2]/select').click()

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[2]/div/div/span[2]/input').click()
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[2]/div/div/span[2]/input').clear()
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[2]/div/div/span[2]/input').click()
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[2]/div/div/span[2]/input').send_keys(
                        data[20])

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                        '2]/div/div[2]/div/div/span[3]/input').send_keys(data[21])
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                        '2]/div/div[3]/div/div/button').click()

                    if data[1] == 4:
                        AS = ['Select value', '8', '9', '10', '11', '13', '14', '15']
                    elif data[1] == 44:
                        AS = ['Select value', '13', '14', '15']
                    elif data[1] == 444:
                        AS = ['Select value', '13', '14']
                    else:
                        AS = ['Select value', '15']

                    dropdown7 = Select(driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[3]'))
                    i = 0
                    c = 0
                    li7 = []
                    for element in dropdown7.options:
                        li7.insert(i, element.text)
                        i += 1
                    for i in range(0, len(AS)):
                        if AS.__getitem__(i) in li7:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'AS Release' is not as Expected",
                                      name='DropdownError_ASRelease')

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                        '2]/div/div[4]/form/div[2]/div[2]/select[3]').click()
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                        '2]/div/div[4]/form/div[2]/div[2]/select[3]').send_keys(data[22])
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                        '2]/div/div[4]/form/div[2]/div[2]/select[3]').click()

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/input[2]').click()
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/input[2]').clear()
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/input[2]').click()
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/input[2]').send_keys(
                        data[23])

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/input[3]').click()
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/input[3]').clear()
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/input[3]').click()
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/input[3]').send_keys(
                        data[24])
                    if data[1] != 444:
                        if 12 > data[22] >= 8:
                            UCT = ['Select value', 'Combined', 'None']
                        elif 12 < data[22] <= 15:
                            UCT = ['Select value', 'Combined', 'Decoupled']
                        else:
                            UCT = ['Select value', 'Combined', 'Decoupled']
                    else:
                        UCT = ['Select value', 'Combined']

                    dropdown7 = Select(driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[2]'))
                    i = 0
                    c = 0
                    li7 = []
                    for element in dropdown7.options:
                        li7.insert(i, element.text)
                        i += 1
                    for i in range(0, len(UCT)):
                        if UCT.__getitem__(i) in li7:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'UE Cat Type' is not as Expected",
                                      name='DropdownError_UECatType')
                    driver.find_element_by_id('attachPDNtype').send_keys(data[25])

                    if data[25].upper() == 'C':
                        if data[1] == 4 or data[1] == 5:
                            UE = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13']
                        elif data[1] == 444:
                            if data[22] == 14:
                                UE = ['0', 'NB-IoT NB1', 'NB-IoT NB2']
                            else:
                                UE = ['0', 'NB-IoT NB1']
                        elif data[1] == 44:
                            UE = ['0', 'CAT-M1']
                        else:
                            UE = []

                        dropdown7 = Select(driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[1]'))
                        i = 0
                        c = 0
                        li7 = []
                        for element in dropdown7.options:
                            li7.insert(i, element.text)
                            i += 1
                        for i in range(0, len(UE)):
                            if UE.__getitem__(i) in li7:
                                assert True
                            else:
                                c += 1
                        if c != 0:
                            allure.attach("Dropdown Values of 'UE Category' is not as Expected",
                                          name='DropdownError_UECategory')

                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                            '2]/div/div[4]/form/div[2]/div[2]/select[1]').click()
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                            '2]/div/div[4]/form/div[2]/div[2]/select[1]').send_keys(data[26])
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                            '2]/div/div[4]/form/div[2]/div[2]/select[1]').click()
                    elif data[25].upper() == 'D':
                        UL = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13']
                        DL = ['0', 'NB-IoT NB1', 'NB-IoT NB2', 'CAT-M1', '1', '2', '3', '4', '5', '6', '7', '8', '9',
                              '10', '11', '12', '13']
                        dropdown7 = Select(driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[1]'))
                        i = 0
                        c = 0
                        li7 = []
                        for element in dropdown7.options:
                            li7.insert(i, element.text)
                            i += 1
                        for i in range(0, len(UL)):
                            if UL.__getitem__(i) in li7:
                                assert True
                            else:
                                c += 1
                        if c != 0:
                            allure.attach("Dropdown Values of 'UL' is not as Expected", name='DropdownError_UL')

                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[1]').click()
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[1]').send_keys(
                            data[27])
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[1]').click()

                        dropdown7 = Select(driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[2]'))
                        i = 0
                        li7 = []
                        c = 0
                        for element in dropdown7.options:
                            li7.insert(i, element.text)
                            i += 1
                        for i in range(0, len(DL)):
                            if DL.__getitem__(i) in li7:
                                assert True
                            else:
                                c += 1
                        if c != 0:
                            allure.attach("Dropdown Values of 'DL' is not as Expected", name='DropdownError_DL')

                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[2]').click()
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[2]').send_keys(
                            data[28])
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[2]').click()

                    if data[29] == 'cell-1':
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                            '2]/div/div[4]/form/div[2]/div[2]/input[1]').click()

                    CQI = ['Auto', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14',
                           '15']

                    dropdown7 = Select(driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[7]'))
                    i = 0
                    c = 0
                    li7 = []
                    for element in dropdown7.options:
                        li7.insert(i, element.text)
                        i += 1
                    for i in range(0, len(CQI)):
                        if CQI.__getitem__(i) in li7:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'CQI' is not as Expected", name='DropdownError_CQI')

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[7]').send_keys(
                        data[30])

                    Algo = ['Milenage', 'XOR', 'TUAK']
                    dropdown7 = Select(driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[9]'))
                    i = 0
                    c = 0
                    li7 = []
                    for element in dropdown7.options:
                        li7.insert(i, element.text)
                        i += 1
                    for i in range(0, len(Algo)):
                        if Algo.__getitem__(i) in li7:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'ALGORITHM' is not as Expected",
                                      name='DropdownError_ALGORITHM')

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[9]').send_keys(
                        data[31])

                    PDN = ['ipv4', 'ipv4v6', 'non-ip']
                    dropdown7 = Select(driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[4]'))
                    i = 0
                    c = 0
                    li7 = []
                    for element in dropdown7.options:
                        li7.insert(i, element.text)
                        i += 1
                    for i in range(0, len(PDN)):
                        if PDN.__getitem__(i) in li7:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'PDN Type' is not as Expected",
                                      name='DropdownError_PDN_Type')

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[4]').click()

                    # driver.find_element_by_xpath('/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[4]').send_keys(data[34])
                    if driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[8]').is_enabled():
                        CA = ['Disable', 'Enable']
                        dropdown7 = Select(driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[8]'))
                        i = 0
                        c = 0
                        li7 = []
                        for element in dropdown7.options:
                            li7.insert(i, element.text)
                            i += 1
                        for i in range(0, len(CA)):
                            if CA.__getitem__(i) in li7:
                                assert True
                            else:
                                c += 1
                        if c != 0:
                            allure.attach("Dropdown Values of 'Combined Attach' is not as Expected",
                                          name='DropdownError_CombinedAttach')
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[8]').send_keys(
                            data[35])

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/input[4]').send_keys(
                        data[36])

                    RI = ['Auto', '1', '2', '3', '4']
                    dropdown7 = Select(driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[5]'))
                    i = 0
                    c = 0
                    li7 = []
                    for element in dropdown7.options:
                        li7.insert(i, element.text)
                        i += 1
                    for i in range(0, len(RI)):
                        if RI.__getitem__(i) in li7:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'RI' is not as Expected", name='DropdownError_RI')

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[5]').send_keys(
                        data[37])

                    PMI = ['Auto', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14',
                           '15']

                    dropdown7 = Select(driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[6]'))
                    i = 0
                    c = 0
                    li7 = []
                    for element in dropdown7.options:
                        li7.insert(i, element.text)
                        i += 1
                    for i in range(0, len(PMI)):
                        if PMI.__getitem__(i) in li7:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'CQI' is not as Expected", name='DropdownError_PMI')

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[6]').send_keys(
                        data[38])
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                        '2]/div/div[4]/form/div[2]/div[2]/select[1]').click()
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                        '2]/div/div[4]/form/div[2]/div[2]/select[1]').send_keys(data[26])
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                        '2]/div/div[4]/form/div[2]/div[2]/select[1]').click()

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[1]/label/span').click()
                    if data[65] != 'N/A' and driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[11]').is_enabled():
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[11]').click()
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[11]').send_keys(
                            data[65])
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[11]').click()
                    if data[66] != 'N/A' and driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[12]').is_enabled():
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[12]').click()
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[12]').send_keys(
                            data[66])
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[12]').click()
                    if data[67] != 'N/A' and driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[13]').is_enabled():
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[13]').click()
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[13]').send_keys(
                            data[67])
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[13]').click()
                    if data[68] != 'N/A' and driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[14]').is_enabled():
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[14]').click()
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[14]').send_keys(
                            data[68])
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[14]').click()
                    if data[69] != 'N/A' and driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[15]').is_enabled():
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[15]').click()
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[15]').send_keys(
                            data[69])
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[15]').click()
                    if data[70] != 'N/A' and driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[7]').is_enabled():
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[7]').click()
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[7]').send_keys(
                            data[70])
                        driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/select[7]').click()

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/button').click()
                    try:
                        A = driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div/div[1]/div/div[1]/h3[1]').is_displayed()
                        B = driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[1]/div/div').is_displayed() or driver.find_element_by_xpath(
                            '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/span').is_displayed()
                    except:
                        None

                    finally:
                        if A or B:
                            assert True
                            if B:
                                SUB = 1
                        else:
                            allure.attach(driver.get_screenshot_as_png(), name='TestSubscriber_Failure',
                                          attachment_type=AttachmentType.PNG)
                            assert False
            except:
                driver.find_element_by_xpath(
                    '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[2]/div[2]/button').click()
                allure.attach(driver.get_screenshot_as_png(), name='TestSubscriber_missingdata',
                              attachment_type=AttachmentType.PNG)
                if driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div[4]/form/div[1]/div/div').is_displayed() or driver.find_element_by_xpath(
                    '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/span').is_displayed():
                    assert True
                    SUB = 1
                else:
                    allure.attach(driver.get_screenshot_as_png(), name='TestSubscriber_Failure',
                                  attachment_type=AttachmentType.PNG)
                    assert False

            if SUB == 0:
                try:
                    P = driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div/div[1]/div/div[2]/span[2]/select').text
                    if data[19].upper() in P.upper():
                        assert True
                    else:
                        allure.attach(
                            "Profile Type Values of UserPlane config dosenot match with the value of Subscriber Config",
                            name='ProfileType_Error')
                        assert False

                    ET = ['No', 'Yes']
                    dropdown8 = Select(driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div/div[1]/div/div[2]/select'))
                    i = 0
                    c = 0
                    li8 = []
                    for element in dropdown8.options:
                        li8.insert(i, element.text)
                        i += 1
                    for i in range(0, len(ET)):
                        if ET.__getitem__(i) in li8:
                            assert True
                        else:
                            c += 1
                    if c != 0:
                        allure.attach("Dropdown Values of 'External Traffic' is not as Expected",
                                      name='DropdownError_DataType')

                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div/div[1]/div/div[2]/select').send_keys(
                        data[39])
                    driver.find_element_by_xpath(
                        '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div/div[1]/div/div[2]/select').click()

                    with allure.step("User Plane"):
                        DT = ['Select', 'UDP', 'VOLTE', 'No Data']
                        dropdown8 = Select(driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[1]'))
                        i = 0
                        c = 0
                        li8 = []
                        for element in dropdown8.options:
                            li8.insert(i, element.text)
                            i += 1
                        for i in range(0, len(DT)):
                            if DT.__getitem__(i) in li8:
                                assert True
                            else:
                                c += 1
                        if c != 0:
                            allure.attach("Dropdown Values of 'Data Type' is not as Expected",
                                          name='DropdownError_DataType')

                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[1]').click()
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[1]').send_keys(data[40])
                        driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[1]').click()
                        if data[40] == 'No Data':
                            a = driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[2]').is_enabled()
                            b = driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[1]').is_enabled()
                            c = driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[2]').is_enabled()
                            d = driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[5]').is_enabled()
                            e = driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[5]').is_enabled()
                            f = driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[10]').is_enabled()
                            g = driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[4]').is_enabled()
                            h = driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[6]').is_enabled()
                            i = driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[7]').is_enabled()

                            if a == b == c == d == e == f == g == h == i == False:
                                assert True
                            else:
                                allure.attach("On selecting VOLTE all the desired fileds are not disabled",
                                              name='VOLTEError')

                            driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/button').click()
                        else:
                            if data[40].upper() == 'UDP' or data[40].upper() == 'FTP':
                                DD = ['Uplink', 'Downlink', 'Both']
                                dropdown8 = Select(
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[2]'))
                                i = 0
                                c = 0
                                li8 = []
                                for element in dropdown8.options:
                                    li8.insert(i, element.text)
                                    i += 1
                                for i in range(0, len(DD)):
                                    if DD.__getitem__(i) in li8:
                                        assert True
                                    else:
                                        c += 1
                                if c != 0:
                                    allure.attach("Dropdown Values of 'Data Type' is not as Expected",
                                                  name='DropdownError_DataDirection')
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[2]').click()
                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[2]').send_keys(
                                    data[41])
                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[2]').click()

                                if data[41] == 'Both':
                                    if driver.find_element_by_xpath(
                                            '//*[@id="det4g1"]/form/div/div/input[1]').is_enabled() and driver.find_element_by_xpath(
                                        '//*[@id="det4g1"]/form/div/div/input[2]').is_enabled():
                                        assert True
                                    else:
                                        allure.attach(
                                            'Selecting Both in Data Direction do not have both UL and DL Bitrate enabled',
                                            name='BitrateError')

                                    # time.sleep(1)
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[3]').click()
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[3]').send_keys(
                                        data[64])
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[3]').click()
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[1]').click()
                                    # time.sleep(1)
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[1]').clear()
                                    # time.sleep(1)
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[1]').send_keys(
                                        str(data[42]))
                                    # time.sleep(1)
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[4]').click()
                                    driver.find_element_by_xpath(
                                        '//*[@id="det4g1"]/form/div/div/select[4]').send_keys(data[64])
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[4]').click()
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[2]').click()
                                    # time.sleep(1)
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[2]').clear()
                                    # time.sleep(1)
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[2]').send_keys(
                                        str(data[43]))
                                elif data[41] == 'Uplink':
                                    a = driver.find_element_by_xpath(
                                        '//*[@id="det4g1"]/form/div/div/input[1]').is_enabled()
                                    if driver.find_element_by_xpath(
                                            '//*[@id="det4g1"]/form/div/div/input[2]').is_enabled() and a == False:
                                        assert True
                                    else:
                                        allure.attach(
                                            'Selecting Uplink in Data Direction do not have correct UL and DL Bitrate enabled',
                                            name='BitrateError')
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[4]').click()
                                    driver.find_element_by_xpath(
                                        '//*[@id="det4g1"]/form/div/div/select[4]').send_keys(data[64])
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[4]').click()
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[2]').click()
                                    # time.sleep(1)
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[2]').clear()
                                    # time.sleep(1)
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[2]').send_keys(
                                        str(data[43]))
                                elif data[41] == 'Downlink':
                                    a = driver.find_element_by_xpath(
                                        '//*[@id="det4g1"]/form/div/div/input[2]').is_enabled()
                                    if driver.find_element_by_xpath(
                                            '//*[@id="det4g1"]/form/div/div/input[1]').is_enabled() and a == False:
                                        assert True
                                    else:
                                        allure.attach(
                                            'Selecting Uplink in Data Direction do not have correct UL and DL Bitrate enabled',
                                            name='BitrateError')
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[3]').click()
                                    driver.find_element_by_xpath(
                                        '//*[@id="det4g1"]/form/div/div/select[3]').send_keys(data[64])
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[3]').click()
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[1]').click()
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[1]').clear()
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[1]').send_keys(
                                        str(data[42]))
                                DP = ['Continuos', 'Intermittent']
                                dropdown8 = Select(
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[5]'))
                                i = 0
                                c = 0
                                li8 = []
                                for element in dropdown8.options:
                                    li8.insert(i, element.text)
                                    i += 1
                                for i in range(0, len(DP)):
                                    if DP.__getitem__(i) in li8:
                                        assert True
                                    else:
                                        c += 1
                                if c != 0:
                                    allure.attach("Dropdown Values of 'Data Type' is not as Expected",
                                                  name='DropdownError_DataPattern')

                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[5]').click()
                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[5]').send_keys(
                                    data[44])
                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[5]').click()
                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[3]').send_keys(
                                    data[45])
                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[5]').click()
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[5]').clear()
                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[5]').send_keys(
                                    data[46])
                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[6]').click()
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[6]').clear()
                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[6]').send_keys(
                                    data[47])
                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[7]').click()
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[7]').clear()
                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[7]').send_keys(
                                    data[48])
                                if data[40] == 'FTP':
                                    a = driver.find_element_by_xpath(
                                        '//*[@id="det4g1"]/form/div/div/input[10]').is_enabled()
                                    if a:
                                        allure.attach("On selecting FTP InterPacket Delay is not disabled",
                                                      name='FTPError')
                                    else:
                                        assert True
                                if data[44] == 'Intermittent' and data[49] != 'NA':
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[10]').click()
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[10]').send_keys(
                                        data[49])
                                if data[50] != 'N/A' and data[40] != 'FTP':
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[4]').click()
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[4]').clear()
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[4]').send_keys(50)

                                if data[51] != 'N/A':
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[9]').click()
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[9]').clear()
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[9]').send_keys(
                                        data[51])

                                S = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/input[8]').get_attribute(
                                    'value')
                                if int(S) == data[21]:
                                    assert True
                                else:
                                    allure.attach(
                                        "# of subscriber Values of UserPlane config dosenot match with the value of Subscriber Config",
                                        name='#ofsubscriber_Error')
                                    assert False

                                IR = ['Range 1']
                                dropdown8 = Select(driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/select[6]'))
                                i = 0
                                c = 0
                                li8 = []
                                for element in dropdown8.options:
                                    li8.insert(i, element.text)
                                    i += 1
                                for i in range(0, len(IR)):
                                    if IR.__getitem__(i) in li8:
                                        assert True
                                    else:
                                        c += 1
                                if c != 0:
                                    allure.attach("Dropdown Values of 'IMSI Range' is not as Expected",
                                                  name='DropdownError_IMSIRange')

                                # driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[6]').send_keys(data[50])
                                # IMSI RANGE
                                if data[40] == 'FTP':
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/div/input').click()
                                    # change the send key for ftp password
                                    driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/div/input').send_keys(
                                        data[71])
                                driver.find_element_by_xpath(
                                    '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div/div[2]/div/div/form/div/div/button').click()

                            else:
                                a = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/select[2]').is_enabled()
                                b = driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[1]').is_enabled()
                                c = driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[2]').is_enabled()
                                d = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/select[5]').is_enabled()
                                e = driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[5]').is_enabled()
                                f = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/input[10]').is_enabled()
                                g = driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[4]').is_enabled()

                                if a == b == c == d == e == f == g == False:
                                    assert True
                                else:
                                    allure.attach("On selecting VOLTE all the desired fileds are not disabled",
                                                  name='VOLTEError')

                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[3]').send_keys(
                                    data[45])
                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[6]').click()
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[6]').clear()
                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[6]').send_keys(
                                    data[47])
                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[7]').click()
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[7]').clear()
                                # time.sleep(1)
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[7]').send_keys(
                                    data[48])
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[9]').click()
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[9]').clear()
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[9]').send_keys(
                                    data[51])

                                S = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/input[8]').get_attribute('value')
                                if int(S) == data[21]:
                                    assert True
                                else:
                                    allure.attach(
                                        "# of subscriber Values of UserPlane config dosenot match with the value of Subscriber Config",
                                        name='#ofsubscriber_Error')
                                    assert False

                                IR = ['Range 1']
                                dropdown8 = Select(driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/select[6]'))
                                i = 0
                                c = 0
                                li8 = []
                                for element in dropdown8.options:
                                    li8.insert(i, element.text)
                                    i += 1
                                for i in range(0, len(IR)):
                                    if IR.__getitem__(i) in li8:
                                        assert True
                                    else:
                                        c += 1
                                if c != 0:
                                    allure.attach("Dropdown Values of 'IMSI Range' is not as Expected",
                                                  name='DropdownError_IMSIRange')

                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/select[6]').send_keys(
                                    data[50])
                                driver.find_element_by_xpath('//*[@id="volteProfile"]').click()
                                driver.find_element_by_xpath('//*[@id="volteProfile"]').send_keys(data[58])
                                driver.find_element_by_xpath('//*[@id="volteProfile"]').click()
                                try:
                                    if data[58] == 'NEW':
                                        driver.find_element_by_xpath('//*[@id="myModal"]/div/div').is_displayed()
                                        driver.find_element_by_xpath(
                                            '//*[@id="myModal"]/div/div/div[2]/input[1]').send_keys(data[57])
                                        driver.find_element_by_xpath(
                                            '//*[@id="myModal"]/div/div/div[2]/input[2]').clear()
                                        driver.find_element_by_xpath(
                                            '//*[@id="myModal"]/div/div/div[2]/input[2]').click()
                                        driver.find_element_by_xpath(
                                            '//*[@id="myModal"]/div/div/div[2]/input[2]').send_keys(data[59])
                                        driver.find_element_by_xpath(
                                            '//*[@id="myModal"]/div/div/div[2]/select[1]').click()
                                        driver.find_element_by_xpath(
                                            '//*[@id="myModal"]/div/div/div[2]/select[1]').send_keys(data[60])
                                        driver.find_element_by_xpath(
                                            '//*[@id="myModal"]/div/div/div[2]/select[1]').click()
                                        driver.find_element_by_xpath(
                                            '//*[@id="myModal"]/div/div/div[2]/input[4]').send_keys(data[61])
                                        if data[61] == 'None':
                                            driver.find_element_by_xpath(
                                                '//*[@id="myModal"]/div/div/div[2]/input[4]').click()
                                            driver.find_element_by_xpath(
                                                '//*[@id="myModal"]/div/div/div[2]/input[4]').clear()
                                            driver.find_element_by_xpath(
                                                '//*[@id="myModal"]/div/div/div[2]/input[4]').click()
                                            time.sleep(2)
                                        driver.find_element_by_xpath(
                                            '//*[@id="myModal"]/div/div/div[2]/select[2]').click()
                                        driver.find_element_by_xpath(
                                            '//*[@id="myModal"]/div/div/div[2]/select[2]').send_keys(data[62])
                                        driver.find_element_by_xpath(
                                            '//*[@id="myModal"]/div/div/div[2]/select[2]').click()
                                        driver.find_element_by_xpath(
                                            '//*[@id="myModal"]/div/div/div[2]/input[5]').click()
                                        driver.find_element_by_xpath(
                                            '//*[@id="myModal"]/div/div/div[2]/input[5]').clear()
                                        driver.find_element_by_xpath(
                                            '//*[@id="myModal"]/div/div/div[2]/input[5]').send_keys(data[63])
                                        if data[63] == 'None':
                                            driver.find_element_by_xpath(
                                                '//*[@id="myModal"]/div/div/div[2]/input[5]').click()
                                            driver.find_element_by_xpath(
                                                '//*[@id="myModal"]/div/div/div[2]/input[5]').clear()
                                            driver.find_element_by_xpath(
                                                '//*[@id="myModal"]/div/div/div[2]/input[5]').click()
                                            time.sleep(2)
                                        driver.find_element_by_xpath('//*[@id="myModal"]/div/div/div[2]/button').click()
                                        time.sleep(2)
                                        driver.find_element_by_xpath('//*[@id="volteProfile"]').click()
                                        driver.find_element_by_xpath('//*[@id="volteProfile"]').send_keys(data[57])
                                        driver.find_element_by_xpath('//*[@id="volteProfile"]').click()
                                    else:
                                        driver.find_element_by_xpath('//*[@id="volteProfile"]').click()
                                    assert True
                                except NoSuchElementException:
                                    allure.attach('IMSI Profile POP UP Error', name='IMSI_Profile_Error')
                                    allure.attach(driver.get_screenshot_as_png(),
                                                  name='IMSI_Profile_Error',
                                                  attachment_type=AttachmentType.PNG)
                                    assert False
                                if data[57] == 'None' or data[59] == 'None' or data[60] == 'None' or data[
                                    61] == 'None' or data[62] == 'None' or data[63] == 'None':
                                    try:
                                        if driver.find_element_by_xpath(
                                                '//*[@id="myModal"]/div/div/div[1]/span').is_displayed():
                                            driver.find_element_by_xpath('//*[@id="closeModal"]').click()
                                            assert True
                                    except NoSuchElementException:
                                        z = 1
                                        assert False

                                driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/button').click()
                            time.sleep(2)
                            try:
                                A = driver.find_element_by_xpath(
                                    '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[2]/div/div/h3').is_displayed()
                            except:
                                B = driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div/div/span').is_displayed() or driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/div[1]/div').is_displayed() or driver.find_element_by_xpath(
                                    '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[1]/span[3]').is_displayed()
                            finally:
                                if A or B:
                                    assert True
                                    if B:
                                        USER = 1
                                else:
                                    allure.attach(driver.get_screenshot_as_png(), name='TestUserPlane_Failure',
                                                  attachment_type=AttachmentType.PNG)
                                    assert False
                except:
                    if z == 1:
                        allure.attach(
                            'IMSI Profile missing data message not displayed and new profile created without filing mandatory data',
                            name='IMSI_Profile_Error')
                        assert False
                    driver.find_element_by_xpath(
                        '//*[@id="det4g1"]/form/div/div/button').click()
                    allure.attach(driver.get_screenshot_as_png(), name='TestUserPlane_missingdata',
                                  attachment_type=AttachmentType.PNG)
                    if driver.find_element_by_xpath(
                            '//*[@id="det4g1"]/form/div/div/span').is_displayed() or driver.find_element_by_xpath(
                        '//*[@id="det4g1"]/form/div[1]/div').is_displayed() or driver.find_element_by_xpath('/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[1]/span[3]').is_displayed():
                        assert True
                        USER = 1
                    else:
                        allure.attach(driver.get_screenshot_as_png(), name='TestUserPlane_Failure',
                                      attachment_type=AttachmentType.PNG)
                        assert False

                if USER == 0:
                    try:
                        with allure.step("Traffic Profile"):

                            AT = ['Select', 'Staggered', 'Bursty']
                            dropdown9 = Select(driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/select'))
                            i = 0
                            c = 0
                            li9 = []
                            for element in dropdown9.options:
                                li9.insert(i, element.text)
                                i += 1
                            for i in range(0, len(AT)):
                                if AT.__getitem__(i) in li9:
                                    assert True
                                else:
                                    c += 1
                            if c != 0:
                                allure.attach("Dropdown Values of 'Attach Type' is not as Expected",
                                              name='DropdownError_AttachType')
                            driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/select').click()
                            driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/select').send_keys(data[53])
                            driver.find_element_by_xpath(
                                '//*[@id="det4g1"]/form/div/div/select').click()

                            if data[53].upper() == 'BURSTY':
                                a = driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[1]').is_enabled()
                                if not a:
                                    assert True
                                else:
                                    allure.attach("On selecting Attach Type as 'Bursty' Attach rate is enabled",
                                                  name='AttachRateError')
                            else:
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[1]').click()
                                driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[1]').send_keys(
                                    data[54])

                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[2]').send_keys(data[55])
                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/input[3]').send_keys(data[56])
                            driver.find_element_by_xpath(
                                '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div['
                                '1]/input').send_keys(data[57])

                            driver.find_element_by_xpath('//*[@id="det4g1"]/form/div/div/div[1]/div/button').click()

                            try:
                                A = driver.find_element_by_xpath(
                                    '/html/body/app-dashboard/div[1]/main/div/app-manage-config/div/div/div[1]/div[1]/span').is_displayed()
                            except:
                                B = driver.find_element_by_xpath(
                                    '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[1]/span[3]').is_displayed() or driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/form/span[3]').is_displayed() or driver.find_element_by_xpath(
                                    '//*[@id="det4g1"]/div/div').is_displayed() or driver.find_element_by_xpath('/html/body/app-dashboard/div[1]/main/div/notifier-container/ul/li/notifier-notification/p').is_displayed()
                            finally:
                                if A or B:

                                    assert True
                                else:
                                    allure.attach(driver.get_screenshot_as_png(), name='TestTrafficProfile_Failure',
                                                  attachment_type=AttachmentType.PNG)
                                    assert False
                    except:
                        driver.find_element_by_xpath(
                            '//*[@id="det4g1"]/form/div/div/div[1]/div/button').click()
                        allure.attach(driver.get_screenshot_as_png(), name='TestTrafficProfile_missingdata',
                                      attachment_type=AttachmentType.PNG)
                        if driver.find_element_by_xpath(
                                '/html/body/app-dashboard/div[1]/main/div/app-create-configuration/div/div[1]/span[3]').is_displayed() or driver.find_element_by_xpath(
                            '//*[@id="det4g1"]/form/span[3]').is_displayed() or driver.find_element_by_xpath(
                            '//*[@id="det4g1"]/div/div').is_displayed():
                            assert True
                        else:
                            allure.attach(driver.get_screenshot_as_png(), name='TestTrafficProfile_Failure',
                                          attachment_type=AttachmentType.PNG)
                            assert False
    finally:
        driver.find_element_by_xpath('/html/body/app-dashboard/app-header/ul/li[2]/a/img').click()
        driver.find_element_by_xpath('/html/body/app-dashboard/app-header/ul/li[2]/div/a[4]').click()
